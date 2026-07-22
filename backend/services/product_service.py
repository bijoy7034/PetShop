"""Excel bulk-upload for products.

The expected sheet columns (case-insensitive, first row is the header):

    name | category | subcategory | description | base_price | discount_price
      | variant_size | variant_weight | variant_color | variant_sku
      | variant_price | variant_discount_price | variant_stock

One row per variant. Rows sharing the same `name` are merged into a single
product; the first row supplies the product-level fields. `category` (and
`subcategory` if given) are resolved by name against existing categories —
unknown categories fail the row rather than being auto-created.
"""
from io import BytesIO
from itertools import product as _cartesian

from openpyxl import load_workbook

from repository.category_repo import CategoryRepository
from repository.product_repo import ProductRepository
from repository.subcategory_repo import SubcategoryRepository


_VARIANT_AXES = ("size", "weight", "color")


def _dedupe(values):
    """Preserve order, drop blanks and dupes."""
    seen = set()
    out = []
    for v in values or ():
        s = str(v).strip() if v is not None else ""
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out


def expand_option_sets(option_sets, base_price):
    """Turn an OptionSet dict into a list of variant dicts.

    Cartesian product across every non-empty axis. If no axes are populated,
    returns a single blank variant (single-SKU product). Every variant
    starts at base_price / stock=0 / sku=None — office fills real prices
    and SKUs via the per-variant PATCH endpoint after creation.
    """
    axes = []
    for key in _VARIANT_AXES:
        values = _dedupe((option_sets or {}).get(key))
        if values:
            axes.append((key, values))

    if not axes:
        return [
            {
                "size": None, "weight": None, "color": None, "sku": None,
                "price": float(base_price), "discount_price": None, "stock": 0,
            }
        ]

    out = []
    for combo in _cartesian(*[values for _, values in axes]):
        v = {
            "size": None, "weight": None, "color": None, "sku": None,
            "price": float(base_price), "discount_price": None, "stock": 0,
        }
        for (key, _), value in zip(axes, combo):
            v[key] = value
        out.append(v)
    return out

_HEADER_ALIASES = {
    # Product name
    "name": "name",
    "product": "name",
    "product name": "name",
    # External client-facing code for the product (distinct from our
    # system-minted PRD-XXXX). Product-level.
    "client product code": "client_product_code",
    "client_product_code": "client_product_code",
    "product code": "client_product_code",
    "product_code": "client_product_code",
    # Taxonomy
    "category": "category",
    "subcategory": "subcategory",
    "sub category": "subcategory",
    "sub-category": "subcategory",
    "description": "description",
    # Product-level attributes
    "brand": "brand",
    "unit": "unit",
    "images": "images",
    "image": "images",
    # Prices
    "base price": "base_price",
    "base_price": "base_price",
    "discount price": "discount_price",
    "discount_price": "discount_price",
    # Variant axes
    "variant name": "variant_name",
    "variant_name": "variant_name",
    "variant size": "variant_size",
    "variant_size": "variant_size",
    "size": "variant_size",
    "variant weight": "variant_weight",
    "variant_weight": "variant_weight",
    "weight": "variant_weight",
    "variant color": "variant_color",
    "variant_color": "variant_color",
    "color": "variant_color",
    "variant sku": "variant_sku",
    "variant_sku": "variant_sku",
    "sku": "variant_sku",
    # Variant image (per-variant override; separate cell from product images)
    "variant image": "variant_image",
    "variant_image": "variant_image",
    # Prices per variant
    "variant price": "variant_price",
    "variant_price": "variant_price",
    "variant discount price": "variant_discount_price",
    "variant_discount_price": "variant_discount_price",
    # Stock + reorder
    "variant stock": "variant_stock",
    "variant_stock": "variant_stock",
    "opening stock": "variant_stock",
    "stock": "variant_stock",
    "reorder level": "reorder_level",
    "reorder_level": "reorder_level",
}


def _to_str_list(v):
    """Parse a cell that holds a list of values — commas or newlines
    both work as separators. Empty cell → empty list."""
    if v is None:
        return []
    s = str(v)
    parts = [p.strip() for chunk in s.replace(",", "\n").split("\n") for p in [chunk]]
    return [p for p in parts if p]


def _normalize_headers(header_row):
    out = {}
    for i, raw in enumerate(header_row):
        if raw is None:
            continue
        key = _HEADER_ALIASES.get(str(raw).strip().lower())
        if key:
            out[key] = i
    return out


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _to_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _resolve_taxonomy(category_name, subcategory_name):
    """Resolve the (category, subcategory) pair from the two sheet columns.

    - If both are supplied: category must exist, subcategory must exist under
      it (name lookup, case-insensitive on subcategory).
    - If only `category` is supplied: subcategory is left None.
    - If only `subcategory` is supplied: reject — subcategory names aren't
      globally unique, so we need a category to disambiguate.
    """
    if not category_name:
        return None, None, "'category' column is required"
    cat = CategoryRepository.by_name(category_name)
    if not cat:
        return None, None, f"Unknown category '{category_name}'"
    if not subcategory_name:
        return cat, None, None
    # Case-insensitive match to be forgiving on Excel input.
    subs, _ = SubcategoryRepository.list(category_id=cat["_id"], limit=500)
    sub = next(
        (s for s in subs if s["name"].lower() == subcategory_name.lower()),
        None,
    )
    if not sub:
        return None, None, (
            f"Unknown subcategory '{subcategory_name}' in category "
            f"'{cat['name']}'"
        )
    return cat, sub, None


def parse_products_workbook(file_bytes):
    """Return (rows, header_error). Each row is a dict with the raw fields
    plus a `_row` number (1-indexed like Excel)."""
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], "The uploaded file is empty."

    headers = _normalize_headers(header_row)
    if "name" not in headers:
        return [], "Header row must contain a 'name' column."
    if "category" not in headers:
        return [], "Header row must contain a 'category' column."
    if "variant_price" not in headers and "base_price" not in headers:
        return [], "Header row must contain 'base_price' or 'variant_price'."

    out = []
    for idx, row in enumerate(rows_iter, start=2):
        if row is None or all(c is None or c == "" for c in row):
            continue
        out.append(
            {
                "_row": idx,
                # Product-level fields (taken from the first row for a
                # given product name; later rows are ignored for these).
                "name": _to_str(_cell(row, headers.get("name"))),
                "client_product_code": _to_str(_cell(row, headers.get("client_product_code"))),
                "category": _to_str(_cell(row, headers.get("category"))),
                "subcategory": _to_str(_cell(row, headers.get("subcategory"))),
                "description": _to_str(_cell(row, headers.get("description"))),
                "brand": _to_str(_cell(row, headers.get("brand"))),
                "unit": _to_str(_cell(row, headers.get("unit"))),
                "images": _to_str_list(_cell(row, headers.get("images"))),
                "base_price": _to_float(_cell(row, headers.get("base_price"))),
                "discount_price": _to_float(_cell(row, headers.get("discount_price"))),
                # Variant-level fields.
                "variant_name": _to_str(_cell(row, headers.get("variant_name"))),
                "variant_size": _to_str(_cell(row, headers.get("variant_size"))),
                "variant_weight": _to_str(_cell(row, headers.get("variant_weight"))),
                "variant_color": _to_str(_cell(row, headers.get("variant_color"))),
                "variant_sku": _to_str(_cell(row, headers.get("variant_sku"))),
                "variant_image": _to_str(_cell(row, headers.get("variant_image"))),
                "variant_price": _to_float(_cell(row, headers.get("variant_price"))),
                "variant_discount_price": _to_float(
                    _cell(row, headers.get("variant_discount_price"))
                ),
                "variant_stock": _to_int(_cell(row, headers.get("variant_stock"))),
                "reorder_level": _to_int(_cell(row, headers.get("reorder_level"))),
            }
        )
    return out, None


def import_products(file_bytes):
    rows, header_error = parse_products_workbook(file_bytes)
    if header_error:
        return {
            "created": 0,
            "updated": 0,
            "failed": 0,
            "rows": [{"row": 1, "action": "header_error", "error": header_error}],
        }

    # Merge rows sharing the same name into a single product with N variants.
    by_name = {}
    order = []
    for r in rows:
        if not r["name"]:
            by_name.setdefault(("__missing__", r["_row"]), r)
            order.append(("__missing__", r["_row"]))
            continue
        key = r["name"]
        if key not in by_name:
            by_name[key] = {"first": r, "variants": []}
            order.append(key)
        by_name[key]["variants"].append(r)

    reports = []
    created = updated = failed = 0

    for key in order:
        if isinstance(key, tuple):
            reports.append(
                {"row": key[1], "action": "failed", "error": "Product name is required"}
            )
            failed += 1
            continue
        bundle = by_name[key]
        head = bundle["first"]

        cat, sub, tax_err = _resolve_taxonomy(head["category"], head["subcategory"])
        if tax_err:
            reports.append(
                {"row": head["_row"], "action": "failed", "product_name": key, "error": tax_err}
            )
            failed += len(bundle["variants"])
            continue
        # Bulk upload requires a subcategory — Product now stores it as a
        # required field. Reject rows that only supplied a category so the
        # sheet can't create half-tagged products.
        if sub is None:
            reports.append(
                {
                    "row": head["_row"],
                    "action": "failed",
                    "product_name": key,
                    "error": "'subcategory' column is required for every product",
                }
            )
            failed += len(bundle["variants"])
            continue

        base_price = head["base_price"]
        if base_price is None:
            # Fall back to the first variant's price so simple sheets (single
            # variant per product) still work without a separate base_price
            # column.
            base_price = head["variant_price"]
        if base_price is None:
            reports.append(
                {
                    "row": head["_row"],
                    "action": "failed",
                    "product_name": key,
                    "error": "base_price (or variant_price) is required",
                }
            )
            failed += len(bundle["variants"])
            continue

        variants_payload = []
        for r in bundle["variants"]:
            price = r["variant_price"] if r["variant_price"] is not None else base_price
            variants_payload.append(
                {
                    "name": r["variant_name"],
                    "size": r["variant_size"],
                    "weight": r["variant_weight"],
                    "color": r["variant_color"],
                    "sku": r["variant_sku"],
                    "image": r["variant_image"],
                    "price": price,
                    "discount_price": r["variant_discount_price"],
                    "initial_stock": r["variant_stock"] or 0,
                    "reorder_level": r["reorder_level"] or 0,
                }
            )

        # Lazy import: inventory_repo depends on product_repo indirectly.
        from repository.inventory_repo import InventoryRepository

        existing = ProductRepository.by_name(key)
        if existing:
            # Update product-level fields and append the new variants. We
            # deliberately do NOT drop variants absent from the sheet — bulk
            # upload is additive; explicit variant delete uses its own route.
            update_patch = {
                "subcategory_id": sub["_id"],
                "subcategory_name": sub["name"],
                "category_id": cat["_id"],
                "category_name": cat["name"],
                "description": head["description"],
                "base_price": base_price,
                "discount_price": head["discount_price"],
            }
            # Only overwrite the new product-level fields when the sheet
            # actually provided a value — an empty column shouldn't clobber
            # whatever's already stored.
            if head["client_product_code"]:
                update_patch["client_product_code"] = head["client_product_code"]
            if head["brand"]:
                update_patch["brand"] = head["brand"]
            if head["unit"]:
                update_patch["unit"] = head["unit"]
            if head["images"]:
                update_patch["images"] = head["images"]
            ProductRepository.update(existing["_id"], update_patch)
            for v in variants_payload:
                _, seed = ProductRepository.add_variant(existing["_id"], v)
                if seed:
                    InventoryRepository.create(
                        product_id=existing["_id"],
                        variant_id=seed["variant_id"],
                        variant_label=seed["variant_label"],
                        product_name=key,
                        quantity_on_hand=seed["initial_stock"],
                        reorder_level=seed["reorder_level"],
                    )
            updated += 1
            for r in bundle["variants"]:
                reports.append(
                    {"row": r["_row"], "action": "updated", "product_name": key}
                )
        else:
            p = ProductRepository.insert(
                name=key,
                subcategory_id=sub["_id"],
                subcategory_name=sub["name"],
                category_id=cat["_id"],
                category_name=cat["name"],
                description=head["description"],
                base_price=base_price,
                discount_price=head["discount_price"],
                variants=variants_payload,
                client_product_code=head["client_product_code"],
                unit=head["unit"],
                images=head["images"] or [],
                brand=head["brand"],
            )
            for seed in p.get("_inventory_seed") or []:
                InventoryRepository.create(
                    product_id=p["_id"],
                    variant_id=seed["variant_id"],
                    variant_label=seed.get("variant_label"),
                    product_name=key,
                    quantity_on_hand=seed["initial_stock"],
                    reorder_level=seed["reorder_level"],
                )
            created += 1
            for r in bundle["variants"]:
                reports.append(
                    {"row": r["_row"], "action": "created", "product_name": key}
                )

    return {"created": created, "updated": updated, "failed": failed, "rows": reports}
